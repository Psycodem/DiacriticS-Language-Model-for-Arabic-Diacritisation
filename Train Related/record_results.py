# -*- coding: utf-8 -*-
"""record_results.py — write a finished run's scores into the tracking surfaces.

    python record_results.py \
        --results-csv sadeed_outputs/Qwen3.5-4B-QLoRA-Tashkeel__results.csv \
        --xlsx "../Models to test & Results.xlsx"

Takes the CSV that eval_sadeed.py produces and:

  1. writes its three rows (MSA / CA / Mean) into the `Ours` table on the
     "Results of test models" sheet, matching the existing Times-12 centred
     styling and extending the table range when it runs out of blank rows;
  2. marks the model Completed on the "Models to test" sheet;
  3. prints a Markdown table ready to paste into README.md, and a JSON blob
     for the Google Sheet / ClickUp mirrors.

Idempotent: re-running with the same Model + Domain Track overwrites those rows
in place instead of appending duplicates — the thing that produced the two
conflicting aya-expanse-8b rows already in the sheet.

Nothing is written unless --write is passed; the default is a dry run that
shows you exactly which cells would change.
"""

import argparse
import json
import os
import shutil
import sys
from datetime import date

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter, range_boundaries

RESULTS_SHEET = "Results of test models"
MODELS_SHEET = "Models to test"
OURS_TABLE = "Ours"

METRIC_COLS = ["DER_ce", "DER_noce", "WER_ce", "WER_noce"]
TRACK_ORDER = ["MSA (Modern)", "CA (Classical)", "Mean (MSA + CA)"]


def parse_args():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--results-csv", required=True)
    p.add_argument("--xlsx", default=os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "Models to test & Results.xlsx"))
    p.add_argument("--model-link", default="", help="HF URL, for the Models-to-test sheet")
    p.add_argument("--model-type", default="LLMs", choices=["LLMs", "Task-Specific"])
    p.add_argument("--assigned-to", default="Mahdi")
    p.add_argument("--notes", default="", help="note cell on the Models-to-test sheet")
    p.add_argument("--write", action="store_true",
                   help="actually modify the workbook (default: dry run)")
    p.add_argument("--no-backup", action="store_true")
    return p.parse_args()


def load_results(path):
    df = pd.read_csv(path)
    missing = [c for c in METRIC_COLS + ["Model", "Domain Track"] if c not in df.columns]
    if missing:
        sys.exit(f"FATAL: {path} is missing columns: {missing}")
    df["Domain Track"] = pd.Categorical(df["Domain Track"], categories=TRACK_ORDER, ordered=True)
    return df.sort_values("Domain Track")


def style_from(ws, template_row, col):
    """Clone the look of an existing data cell so appended rows don't stand out."""
    src = ws.cell(row=template_row, column=col)
    return (Font(name=src.font.name or "Times", size=src.font.size or 12,
                 bold=src.font.bold),
            Alignment(horizontal=src.alignment.horizontal or "center",
                      vertical=src.alignment.vertical))


def find_row(ws, min_row, max_row, model, track):
    """Existing row for this Model (col E) + Domain Track (col F), if any."""
    for r in range(min_row, max_row + 1):
        if (str(ws.cell(row=r, column=5).value or "").strip() == model
                and str(ws.cell(row=r, column=6).value or "").strip() == track):
            return r
    return None


def is_blank(ws, r, occupied):
    return (r not in occupied
            and all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, 7)))


def last_row_of_track(ws, min_row, max_row, track, occupied):
    """Bottom row of the block for this Domain Track.

    The sheet is laid out in three contiguous blocks — Mean, then MSA, then CA —
    each followed by blank separator rows. Appending must land at the end of the
    matching block, not in whatever gap happens to come first.
    """
    last = None
    for r in range(min_row, max_row + 1):
        if str(ws.cell(row=r, column=6).value or "").strip() == track or occupied.get(r) == track:
            last = r
    return last


def write_results_sheet(wb, df, dry):
    ws = wb[RESULTS_SHEET]
    table = ws.tables[OURS_TABLE]
    _, y1, _, y2 = range_boundaries(table.ref)     # A1:F34 -> (1,1,6,34)
    other = {n: ws.tables[n] for n in ws.tables if n != OURS_TABLE}
    changes = []
    # Rows this call has already claimed. Without it a dry run — which writes
    # nothing — would hand every incoming row the same target.
    occupied = {}

    for _, row in df.iterrows():
        model, track = str(row["Model"]), str(row["Domain Track"])
        target = find_row(ws, y1 + 1, y2, model, track)
        action = "update"

        if target is None:
            last = last_row_of_track(ws, y1 + 1, y2, track, occupied)
            candidate = (last + 1) if last else y2 + 1

            if candidate <= y2 and is_blank(ws, candidate, occupied):
                target, action = candidate, "fill-gap"
            else:
                # No gap left under this block — insert a row and push the rest
                # of the sheet down. Safe for the Sadeed_Paper table (J1:N11) as
                # long as we never insert above row 12; the Mean block ends at 11.
                target, action = candidate, "insert"
                if not dry:
                    ws.insert_rows(candidate)
                    for name, t in other.items():
                        ox1, oy1, ox2, oy2 = range_boundaries(t.ref)
                        if oy2 >= candidate:      # only shift tables below the cut
                            t.ref = (f"{get_column_letter(ox1)}{oy1 + 1}:"
                                     f"{get_column_letter(ox2)}{oy2 + 1}")
                occupied = {(r + 1 if r >= candidate else r): v
                            for r, v in occupied.items()}
                y2 += 1

        occupied[target] = track
        values = [row[c] for c in METRIC_COLS] + [model, track]
        for ci, val in enumerate(values, start=1):
            font, align = style_from(ws, y1 + 1, ci)
            changes.append((f"{get_column_letter(ci)}{target}",
                            ws.cell(row=target, column=ci).value, val, action))
            if not dry:
                cell = ws.cell(row=target, column=ci)
                cell.value = round(float(val), 2) if ci <= 4 else val
                cell.font, cell.alignment = font, align

    new_ref = f"A{y1}:F{y2}"
    if new_ref != table.ref:
        changes.append((f"table:{OURS_TABLE}", table.ref, new_ref, "extend-range"))
        if not dry:
            table.ref = new_ref
    return changes


def write_models_sheet(wb, model, args, dry):
    """Mark the model Completed on the tracking sheet (add the row if new)."""
    ws = wb[MODELS_SHEET]
    tname = next(iter(ws.tables))
    table = ws.tables[tname]
    x1, y1, x2, y2 = range_boundaries(table.ref)
    changes = []

    target = None
    for r in range(y1 + 1, y2 + 1):
        if str(ws.cell(row=r, column=6).value or "").strip() == model:
            target = r
            break

    if target is None:
        target = next((r for r in range(y1 + 1, y2 + 1) if is_blank(ws, r, {})), None)
        if target is None:
            y2 += 1
            target = y2
            changes.append((f"table:{tname}", table.ref, f"A{y1}:F{y2}", "extend-range"))
            if not dry:
                table.ref = f"A{y1}:F{y2}"

    font, align = style_from(ws, y1 + 1, 2)
    values = {1: args.notes, 2: "Completed", 3: args.model_link,
              4: args.assigned_to, 5: args.model_type, 6: model}
    for ci, val in values.items():
        if val == "" and ws.cell(row=target, column=ci).value:
            continue                       # don't blank an existing note
        changes.append((f"{get_column_letter(ci)}{target}",
                        ws.cell(row=target, column=ci).value, val, "models-sheet"))
        if not dry:
            cell = ws.cell(row=target, column=ci)
            cell.value = val
            cell.font, cell.alignment = font, align
    return changes


def markdown_table(df):
    mean = df[df["Domain Track"] == "Mean (MSA + CA)"]
    src = mean if not mean.empty else df
    lines = ["| Model | WER w/ CE (%) | DER w/ CE (%) | WER w/o CE (%) | DER w/o CE (%) |",
             "|---|---|---|---|---|"]
    for _, r in src.iterrows():
        lines.append(f"| {r['Model']} | {r['WER_ce']} | {r['DER_ce']} | "
                     f"{r['WER_noce']} | {r['DER_noce']} |")
    return "\n".join(lines)


def main():
    args = parse_args()
    xlsx = os.path.abspath(args.xlsx)
    if not os.path.exists(xlsx):
        sys.exit(f"FATAL: workbook not found: {xlsx}")

    df = load_results(args.results_csv)
    model = str(df["Model"].iloc[0])
    dry = not args.write

    print(f"{'DRY RUN — nothing written' if dry else 'WRITING'}")
    print(f"  workbook : {xlsx}")
    print(f"  model    : {model}")
    print(f"  rows     : {len(df)}\n")
    print(df[METRIC_COLS + ["Model", "Domain Track"]].to_string(index=False), "\n")

    if not dry and not args.no_backup:
        bak = f"{xlsx}.{date.today():%Y%m%d}.bak"
        shutil.copy2(xlsx, bak)
        print(f"  backup   : {bak}\n")

    wb = load_workbook(xlsx)
    changes = write_results_sheet(wb, df, dry)
    changes += write_models_sheet(wb, model, args, dry)

    print(f"{'Would change' if dry else 'Changed'} {len(changes)} cells:")
    for coord, old, new, action in changes:
        print(f"  {coord:16s} {str(old)[:28]:30s} -> {str(new)[:28]:30s} [{action}]")

    if not dry:
        wb.save(xlsx)
        print(f"\nsaved -> {xlsx}")

    print("\n--- README table row(s) ---")
    print(markdown_table(df))

    mirror = os.path.splitext(args.results_csv)[0] + "__mirror.json"
    payload = {
        "model": model,
        "recorded": date.today().isoformat(),
        "rows": json.loads(
            df[METRIC_COLS + ["Model", "Domain Track"]].to_json(orient="records")),
    }
    with open(mirror, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"\nmirror payload for Google Sheets / ClickUp -> {mirror}")


if __name__ == "__main__":
    main()
